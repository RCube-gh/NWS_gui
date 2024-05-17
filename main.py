import sys
import os
import tkinter as tk
from tkinter import filedialog
from pathlib import Path
import tkinter.font as font
import datetime
import openpyxl
import time



root=tk.Tk()
root.geometry("640x380")

date=str(datetime.date.today())
date2=int(date.replace('-',''))


def close(event):
    root.withdraw()
    sys.exit()
    
def writepath():
    path2list=filedialog.askopenfilename(initialdir=Path.cwd())
    f=open('path2.txt','w',encoding='UTF-8')
    f.write(path2list)
    f.close()

def revert_btntxt():
    button["text"]="SAVE"
    button["fg"]="black"

def savedata(event=None):
    text1=t1.get()
    text2=t2.get('1.0',tk.END)
    if os.path.isfile(file_path) and os.path.isfile(path):
        wb=openpyxl.load_workbook(path)
        ws=wb.worksheets[0]
        targetline=ws['C1'].value
        ws['A'+str(targetline)].value=date2
        ws['B'+str(targetline)].value=text1
        ws['C'+str(targetline)].value=text2
        button["text"]="SAVED!!"
        button["fg"]="blue"
        ws['C1'].value=targetline+1
        wb.save(path)
        t1.delete(0,tk.END)
        t2.delete('1.0',tk.END)
        t1.focus_set()
        root.after(1000,revert_btntxt)
    else:
        button["text"]="failed"
        button["fg"]="red"
        root.after(1000,revert_btntxt)




file_path='./path2.txt'
path=''

if os.path.isfile(file_path) == False:
    writepath()
else:
    f=open('./path2.txt','r')
    path=f.readline()
    if os.path.isfile(path.rstrip('\n')) == False:
        writepath()
    f.close()
    
root.title(u"[NewWordsStack] PATH="+path)

font1=tk.font.Font(family="Gothic",underline=False,slant="roman",size=30)
dlabel=tk.Label(text=date,font=("Arial",12),background="#888")
label1=tk.Label(text=u'New Word',font="System",background="#888")
label2=tk.Label(text=u'Description',font="System",background="#888")
t1=tk.Entry(root,font=font1,width=20,justify=tk.CENTER)
t2=tk.Text(root,height=7,width=40,font=("Gothic",15))
button=tk.Button(root,text='SAVE',width=20,height=7,font=("System",15,"bold"),background=("#dfdfdf"),command=savedata)

dlabel.pack(anchor=tk.W,ipadx=20,ipady=5)
label1.pack()
t1.pack()
label2.pack()
t2.pack()
button.pack(anchor=tk.S,pady=20)


root.configure(bg="#888")
root.bind('<Escape>', close)
root.bind('<Alt-Return>', savedata)
root.resizable(False,False)
root.mainloop()
